"""
convert_to_json.py
Konversi file Excel "View Individu" harian menjadi data.js
untuk dashboard Akuisisi Board.

Cara pakai:
    python convert_to_json.py "View_Individu_28062026.xlsx"

Hasil:
    data.js  (timpa file lama, siap dipakai dashboard)

Catatan:
- Urutan & nama kolom Excel harus tetap sama seperti file asli
  (A:Unit Kerja, B:Nip, C:Nama, D:#Cif Akuisisi, ... O:Cek pegawai area)
- Jalankan script ini setiap hari setelah file Excel terbaru di-update,
  lalu replace data.js di folder dashboard (folder yang sama dengan index.html).
"""

import sys
import json
import os
import openpyxl

def convert(excel_path, output_path="data.js"):
    if not os.path.exists(excel_path):
        print(f"File tidak ditemukan: {excel_path}")
        sys.exit(1)

    print(f"Membaca {excel_path} ...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    data = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, 16)]  # kolom A..O
        if row[1] is None:  # skip baris kosong (NIP kosong)
            continue
        # bulatkan angka desimal supaya file lebih ringkas
        for i in range(4, 13):
            if isinstance(row[i], float):
                row[i] = round(row[i], 4)
        data.append(row)

    print(f"Total baris valid: {len(data)}")

    j = json.dumps(data, ensure_ascii=False, separators=(",", ":"), default=str)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("const DATA=" + j + ";")

    size_mb = os.path.getsize(output_path) / 1024 / 1024
    print(f"Selesai. {output_path} dibuat ({size_mb:.2f} MB)")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Pemakaian: python convert_to_json.py <file_excel.xlsx> [output.js]")
        sys.exit(1)
    excel_file = sys.argv[1]
    out_file = sys.argv[2] if len(sys.argv) > 2 else "data.js"
    convert(excel_file, out_file)